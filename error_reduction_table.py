"""
ADRAU-LLM  —  Table 1 only (npj academic three-line format)
• Rate (%) and n are merged into a single row  e.g.  "30.3%  (n=12,307)"
• Overuse / Underuse keep separate rows (different denominators)
• Bottom border, footnote, clean cell merging for Task groups
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), 'ADRAU_Table1.xlsx')

wb = Workbook()
ws = wb.active
ws.title = "Table 1"

# ── helpers ───────────────────────────────────────────────────────────────
def S(style='thin', color='000000'):
    return Side(border_style=style, color=color)

def border_row(top=None, bottom=None):
    t = S('medium') if top    else Side(border_style=None)
    b = S('medium') if bottom else Side(border_style=None)
    return Border(top=t, bottom=b)

def apply_row_border(ws, row, ncols, top=False, bottom=False):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        existing = c.border
        c.border = Border(
            top    = S('medium') if top    else existing.top,
            bottom = S('medium') if bottom else existing.bottom,
            left   = existing.left,
            right  = existing.right,
        )

NCOLS = 6

# ── Fonts ────────────────────────────────────────────────────────────────
F_TITLE = Font(name='Arial', size=9,   bold=True)
F_HDR   = Font(name='Arial', size=8.5, bold=True)
F_TASK  = Font(name='Arial', size=8.5, bold=True,  color='000000')
F_BODY  = Font(name='Arial', size=8.5)
F_BLUE  = Font(name='Arial', size=8.5, bold=True,  color='1F4E79')   # base model
F_AMBER = Font(name='Arial', size=8.5, bold=True,  color='833C00')   # physicians
F_TEAL  = Font(name='Arial', size=8.5, bold=True,  color='1D6A45')   # ADRAU-LLM
F_RED   = Font(name='Arial', size=8.5, italic=True, color='8B2020')  # reduction
F_FOOT  = Font(name='Arial', size=7.5, color='555555', italic=True)
F_DASH  = Font(name='Arial', size=8.5, color='999999')

STRIPE = PatternFill('solid', fgColor='F7F7F7')

# ── Column widths ─────────────────────────────────────────────────────────
widths = [20, 34, 22, 22, 22, 36]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════
# ROW 1  — Table title
# ══════════════════════════════════════════════════════════════════════════
ws.row_dimensions[1].height = 20
ws.merge_cells('A1:F1')
c = ws['A1']
c.value     = 'Table 1  |  Performance comparison of ADRAU-LLM, base model, and physicians'
c.font      = F_TITLE
c.alignment = Alignment(horizontal='left', vertical='center')

# ══════════════════════════════════════════════════════════════════════════
# ROW 2  — Column headers  (top + bottom medium border = "top rule")
# ══════════════════════════════════════════════════════════════════════════
ws.row_dimensions[2].height = 22
HEADERS = ['Task', 'Metric', 'Base model', 'Physicians', 'ADRAU-LLM', 'Error reduction']
HALIGN  = ['left', 'left', 'center', 'center', 'center', 'left']
for col, (h, a) in enumerate(zip(HEADERS, HALIGN), 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font      = F_HDR
    c.alignment = Alignment(horizontal=a, vertical='center')
    c.border    = Border(top=S('medium'), bottom=S('medium'))

# ══════════════════════════════════════════════════════════════════════════
# DATA  — merged n + rate into single cell
# ══════════════════════════════════════════════════════════════════════════
#
# Schema per row:
#   task          : str  ('' for continuation rows within same group)
#   metric        : str
#   base_val      : str  (displayed value; '—' = not applicable)
#   phys_val      : str
#   adrau_val     : str
#   reduction     : str
#   group_start   : bool  (True = first row of a new task group)
#   stripe        : bool
#
ROWS = [
    # ── Diagnosis ─────────────────────────────────────────────────────
    ('Diagnosis\n(n = 40,559)',
     'Top-1 error rate\n(error count)',
     '81.1%\n(n = 32,897)',
     '—',
     '30.3%\n(n = 12,307)',
     '−62.6% (count)\n−50.8 pp (rate)\nvs. base model',
     True, True),

    ('',
     'Top-3 error rate\n(error count)',
     '69.9%\n(n = 28,355)',
     '—',
     '24.1%\n(n = 9,764)',
     '−65.6% (count)\n−45.8 pp (rate)\nvs. base model',
     False, False),

    # ── Antibiotic prescribing ────────────────────────────────────────
    ('Antibiotic\nprescribing\n(n = 31,777)',
     'Overall error rate\n(error count)',
     '15.2%\n(n = 4,825)',
     '39.6%\n(n = 12,581)',
     '13.4%\n(n = 4,250)',
     '−11.9% vs. base model\n−66.2% vs. physicians\n(−26.2 pp rate)',
     True, True),

    ('',
     'Overuse error count\n("Never" appropriate)',
     '—',
     'n = 12,273',
     'n = 3,922',
     '−68.0% vs. physicians',
     False, False),

    ('',
     'Underuse error count\n("Always" appropriate)',
     'n = 328',
     '—',
     'n = 638',
     '↑94.5% vs. base model†',
     False, True),
]

R0 = 3   # first data row

for i, (task, metric, base, phys, adrau, reduc,
        grp_start, stripe) in enumerate(ROWS):

    r = R0 + i
    # row height: taller for multi-line cells
    ws.row_dimensions[r].height = 42 if '\n' in metric else 22

    fill = STRIPE if stripe else None

    # separator line above new group (not the very first)
    top_border = grp_start and i > 0

    # ── col 1: Task ─────────────────────────────────────────────────
    c = ws.cell(row=r, column=1, value=task)
    c.font      = F_TASK
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if fill:       c.fill   = fill
    if top_border: c.border = Border(top=S('medium'))

    # ── col 2: Metric ────────────────────────────────────────────────
    c = ws.cell(row=r, column=2, value=metric)
    c.font      = F_BODY
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if fill:       c.fill   = fill
    if top_border: c.border = Border(top=S('medium'))

    # ── col 3: Base model ────────────────────────────────────────────
    c = ws.cell(row=r, column=3, value=base)
    c.font      = F_DASH if base == '—' else F_BLUE
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if fill:       c.fill   = fill
    if top_border: c.border = Border(top=S('medium'))

    # ── col 4: Physicians ────────────────────────────────────────────
    c = ws.cell(row=r, column=4, value=phys)
    c.font      = F_DASH if phys == '—' else F_AMBER
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if fill:       c.fill   = fill
    if top_border: c.border = Border(top=S('medium'))

    # ── col 5: ADRAU-LLM ────────────────────────────────────────────
    c = ws.cell(row=r, column=5, value=adrau)
    c.font      = F_DASH if adrau == '—' else F_TEAL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if fill:       c.fill   = fill
    if top_border: c.border = Border(top=S('medium'))

    # ── col 6: Reduction ────────────────────────────────────────────
    c = ws.cell(row=r, column=6, value=reduc)
    c.font      = F_RED
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    if fill:       c.fill   = fill
    if top_border: c.border = Border(top=S('medium'))

# ── Bottom rule ───────────────────────────────────────────────────────────
last_r = R0 + len(ROWS) - 1
apply_row_border(ws, last_r, NCOLS, bottom=True)

# ── Task cell merging within groups ──────────────────────────────────────
# Diagnosis: rows R0 to R0+1 (2 rows)
diag_start, diag_end = R0, R0 + 1
ws.merge_cells(f'A{diag_start}:A{diag_end}')
# Antibiotic: rows R0+2 to R0+4 (3 rows)
anti_start, anti_end = R0 + 2, R0 + 4
ws.merge_cells(f'A{anti_start}:A{anti_end}')
# Re-apply font/align after merge (openpyxl resets on merge)
for rng, fill_it in [(f'A{diag_start}', True), (f'A{anti_start}', True)]:
    c = ws[rng]
    c.font      = F_TASK
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill      = STRIPE

# ══════════════════════════════════════════════════════════════════════════
# Footnote
# ══════════════════════════════════════════════════════════════════════════
fn = last_r + 2
ws.row_dimensions[fn].height = 36
ws.merge_cells(f'A{fn}:F{fn}')
c = ws.cell(row=fn, column=1,
    value=(
        '† Underuse (cases where antibiotic is "Always" appropriate) increased in ADRAU-LLM vs. base model, '
        'reflecting more aggressive treatment in this subset; this is considered a separate safety concern from overuse. '
        'pp, percentage points; n, number of cases/errors; —, comparator not applicable for this metric.'
    ))
c.font      = F_FOOT
c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ══════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════
wb.save(OUT)
print(f"✓ Table saved → {OUT}")